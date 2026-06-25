import re
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


PDF_PATH = "/Users/sankarsangili/Downloads/gpay_statement_20251001_20260331.pdf"
OUTPUT_PATH = "/Users/sankarsangili/Downloads/gpay_transactions.xlsx"


def parse_amount(amount_str):
    return float(amount_str.replace("₹", "").replace(",", ""))


def extract_transactions(pdf_path):
    transactions = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue

            lines = [l.strip() for l in text.split("\n") if l.strip()]
            i = 0
            while i < len(lines):
                line = lines[i]

                # Match: "02Oct,2025 PaidtoBSNLBILLDESK ₹997"
                m = re.match(
                    r"^(\d{2}[A-Za-z]{3},\d{4})\s+(.+?)\s+(₹[\d,]+(?:\.\d{2})?)$",
                    line,
                )
                if m:
                    date_str, details, amount_str = m.group(1), m.group(2), m.group(3)

                    if details.startswith("Paidto"):
                        txn_type = "Sent"
                        counterparty = details[len("Paidto"):]
                    elif details.startswith("Receivedfrom"):
                        txn_type = "Received"
                        counterparty = details[len("Receivedfrom"):]
                    else:
                        txn_type = "Other"
                        counterparty = details

                    time_str = upi_id = account = ""

                    if i + 1 < len(lines):
                        tm = re.match(
                            r"^(\d{2}:\d{2}[AP]M)\s+UPITransactionID:(\d+)$",
                            lines[i + 1],
                        )
                        if tm:
                            time_str = tm.group(1)
                            upi_id = tm.group(2)
                            i += 1

                    if i + 1 < len(lines) and re.match(
                        r"^(Paidby|Paidto)StateBankofIndia", lines[i + 1]
                    ):
                        account = lines[i + 1]
                        i += 1

                    try:
                        date_obj = datetime.strptime(date_str, "%d%b,%Y").strftime("%d-%b-%Y")
                    except ValueError:
                        date_obj = date_str

                    transactions.append(
                        {
                            "Date": date_obj,
                            "Time": time_str,
                            "Type": txn_type,
                            "Counterparty": counterparty,
                            "Amount (₹)": parse_amount(amount_str),
                            "UPI Transaction ID": upi_id,
                            "Account": account,
                        }
                    )

                i += 1

    return transactions


def write_excel(transactions, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = ["Date", "Time", "Type", "Counterparty", "Amount (₹)", "UPI Transaction ID", "Account"]

    # Header style
    header_fill = PatternFill("solid", fgColor="1A73E8")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    ws.row_dimensions[1].height = 20

    # Row styles
    sent_fill = PatternFill("solid", fgColor="FFF0F0")
    recv_fill = PatternFill("solid", fgColor="F0FFF0")

    for row_idx, txn in enumerate(transactions, 2):
        fill = sent_fill if txn["Type"] == "Sent" else recv_fill
        for col, key in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col, value=txn[key])
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if key == "Amount (₹)":
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif key == "Type":
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(
                    bold=True,
                    color="CC0000" if txn["Type"] == "Sent" else "007700",
                )

    # Summary rows
    summary_row = len(transactions) + 3
    total_sent = sum(t["Amount (₹)"] for t in transactions if t["Type"] == "Sent")
    total_recv = sum(t["Amount (₹)"] for t in transactions if t["Type"] == "Received")

    summary_font = Font(bold=True, size=11)
    for label, value, color in [
        ("Total Sent", total_sent, "CC0000"),
        ("Total Received", total_recv, "007700"),
        ("Net", total_recv - total_sent, "1A73E8"),
    ]:
        ws.cell(row=summary_row, column=4, value=label).font = summary_font
        cell = ws.cell(row=summary_row, column=5, value=value)
        cell.font = Font(bold=True, color=color, size=11)
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal="right")
        summary_row += 1

    # Column widths
    col_widths = [13, 10, 10, 35, 14, 22, 30]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    wb.save(output_path)
    return len(transactions), total_sent, total_recv


if __name__ == "__main__":
    print("Extracting transactions...")
    transactions = extract_transactions(PDF_PATH)
    print(f"Found {len(transactions)} transactions")

    count, sent, recv = write_excel(transactions, OUTPUT_PATH)
    print(f"\nSaved to: {OUTPUT_PATH}")
    print(f"  Transactions : {count}")
    print(f"  Total Sent   : ₹{sent:,.2f}")
    print(f"  Total Received: ₹{recv:,.2f}")
    print(f"  Net          : ₹{recv - sent:,.2f}")
